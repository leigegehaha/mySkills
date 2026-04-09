import json
import re
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

base = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.cwd()
manifest = json.loads((base / 'invoice_manifest_unread_pdf_only.json').read_text())

rows = []
for item in manifest:
    subject = item.get('subject', '') or ''
    preview = item.get('body_preview', '') or ''
    text = f"{subject} {preview}"

    number_match = re.search(r'发票号码[:：]?\s*([0-9A-Za-z]+)', text)
    amount_match = re.search(r'(?:价税合计(?:金额)?|发票金额|合计)[:：]?\s*[￥¥]?\s*([0-9]+(?:\.[0-9]{1,2})?)', text)
    buyer_match = re.search(r'购方名称[:：]?\s*([^\s，,<>]+(?:（[^）]+）)?[^\s，,<>]*)', text)
    seller_match = re.search(r'(?:来自【([^】]+)】|销方名称[:：]?\s*([^\s，,<>]+(?:（[^）]+）)?[^\s，,<>]*)|收到一张【([^】]+)】开具)', text)

    seller = ''
    if seller_match:
        seller = next((g for g in seller_match.groups() if g), '')

    rows.append({
        '日期': item.get('date', ''),
        '发票号码': number_match.group(1) if number_match else '',
        '金额': float(amount_match.group(1)) if amount_match else None,
        '购方名称': buyer_match.group(1) if buyer_match else '',
        '销方名称': seller,
        '主题': item.get('subject', ''),
        '状态': item.get('status', ''),
        '已下载PDF': item.get('saved_files', ''),
        '手动下载链接': item.get('manual_download_link', ''),
        'UID': item.get('uid', ''),
        '备注': '邮件内已有PDF附件，已下载' if item.get('status') == '已下载' else '需手动打开链接下载PDF',
    })

rows.sort(key=lambda r: (r['状态'], r['日期']))

wb = Workbook()
ws = wb.active
ws.title = '未读发票PDF汇总'
headers = ['日期', '发票号码', '金额', '购方名称', '销方名称', '主题', '状态', '已下载PDF', '手动下载链接', 'UID', '备注']
ws.append(headers)

header_fill = PatternFill('solid', fgColor='FFF3E0')
header_font = Font(name='Arial', bold=True, color='E65100')
body_font = Font(name='Arial')
manual_fill = PatternFill('solid', fgColor='FFF59D')

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

for row_data in rows:
    ws.append([row_data[h] for h in headers])

for row in ws.iter_rows(min_row=2):
    status = row[6].value
    for cell in row:
        cell.font = body_font
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        if status == '需手动下载':
            cell.fill = manual_fill

for cell in ws['C'][1:]:
    cell.number_format = '#,##0.00'

widths = {
    'A': 22, 'B': 24, 'C': 12, 'D': 28, 'E': 28, 'F': 50,
    'G': 14, 'H': 60, 'I': 60, 'J': 10, 'K': 26,
}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

ws.freeze_panes = 'A2'
output = base / 'qq邮箱未读发票PDF汇总.xlsx'
wb.save(output)
print(output)
