import json
import re
import subprocess
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

base = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.cwd()
manifest_path = base / 'invoice_manifest_unread_pdf_only.json'
excel_path = base / 'qq邮箱未读发票PDF汇总.xlsx'

rows = json.loads(manifest_path.read_text()) if manifest_path.exists() else []


def extract_text(pdf_path: Path) -> str:
    cmds = [
        ['mdls', '-name', 'kMDItemTextContent', str(pdf_path)],
        ['strings', str(pdf_path)],
    ]
    for cmd in cmds:
        try:
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=20)
            if result.returncode == 0 and result.stdout.strip():
                text = result.stdout
                if 'kMDItemTextContent' in text:
                    text = text.split('=', 1)[-1].strip().strip('()').strip()
                if len(text.strip()) > 20:
                    return text
        except Exception:
            pass
    return ''


def parse_info(text: str):
    compact = re.sub(r'\s+', ' ', text)
    amount = ''
    seller = ''
    issue_date = ''

    amount_patterns = [
        r'价税合计[（(]小写[)）]?[:：]?\s*[¥￥]?\s*([0-9]+(?:\.[0-9]{1,2})?)',
        r'价税合计[:：]?\s*[¥￥]?\s*([0-9]+(?:\.[0-9]{1,2})?)',
        r'合计[:：]?\s*[¥￥]?\s*([0-9]+(?:\.[0-9]{1,2})?)',
        r'小写[:：]?\s*[¥￥]?\s*([0-9]+(?:\.[0-9]{1,2})?)',
    ]
    for p in amount_patterns:
        m = re.search(p, compact, re.I)
        if m:
            amount = m.group(1)
            break

    seller_patterns = [
        r'销售方名称[:：]?\s*([^\n]+)',
        r'销方名称[:：]?\s*([^\n]+)',
        r'名称[:：]?\s*([^\n]{4,40}公司[^\n]*)',
    ]
    for p in seller_patterns:
        m = re.search(p, text, re.I)
        if m:
            seller = re.sub(r'\s+', ' ', m.group(1)).strip()
            seller = re.split(r'纳税人识别号|统一社会信用代码|地址', seller)[0].strip()
            break

    date_patterns = [
        r'开票日期[:：]?\s*([0-9]{4}[-/.年][0-9]{1,2}[-/.月][0-9]{1,2}日?)',
        r'日期[:：]?\s*([0-9]{4}[-/.年][0-9]{1,2}[-/.月][0-9]{1,2}日?)',
    ]
    for p in date_patterns:
        m = re.search(p, compact)
        if m:
            issue_date = m.group(1)
            break

    summary_parts = []
    if seller:
        summary_parts.append(f'开票方为{seller}')
    if issue_date:
        summary_parts.append(f'开票日期为{issue_date}')
    if amount:
        summary_parts.append(f'金额为{amount}元')
    if not summary_parts:
        summary_parts.append('已读取PDF，但关键信息提取不完整，建议人工复核原票面')
    summary = '；'.join(summary_parts) + '。'
    return amount, seller, issue_date, summary, compact[:1000]

wb = load_workbook(excel_path)
if 'PDF提取汇总' in wb.sheetnames:
    del wb['PDF提取汇总']
ws = wb.create_sheet('PDF提取汇总')
headers = ['UID', '状态', 'PDF路径', 'PDF提取金额', 'PDF提取开票方', 'PDF提取开票日期', 'AI总结', '提取文本片段']
ws.append(headers)

header_fill = PatternFill('solid', fgColor='E3F2FD')
header_font = Font(name='Arial', bold=True, color='0D47A1')
manual_fill = PatternFill('solid', fgColor='FFF59D')
body_font = Font(name='Arial')

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

for item in rows:
    status = item.get('status', '')
    files = [p.strip() for p in (item.get('saved_files', '') or '').split('|') if p.strip()]
    if files:
        for file_path in files:
            pdf_path = Path(file_path)
            text = extract_text(pdf_path) if pdf_path.exists() else ''
            amount, seller, issue_date, summary, snippet = parse_info(text)
            ws.append([item.get('uid', ''), status, file_path, amount, seller, issue_date, summary, snippet])
    else:
        ws.append([item.get('uid', ''), status, '', '', '', '', '该发票未下载PDF，需要手动下载后再提取。', ''])

for row in ws.iter_rows(min_row=2):
    status = row[1].value
    for cell in row:
        cell.font = body_font
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        if status == '需手动下载':
            cell.fill = manual_fill

widths = {'A': 10, 'B': 14, 'C': 60, 'D': 14, 'E': 28, 'F': 18, 'G': 50, 'H': 80}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

ws.freeze_panes = 'A2'
wb.save(excel_path)
print(excel_path)
